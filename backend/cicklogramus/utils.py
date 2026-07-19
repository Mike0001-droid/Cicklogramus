from openpyxl import Workbook
from datetime import datetime
from .models import Project, Task
from django.http import HttpResponse
from .exceptions import NoTaskException
from openpyxl.utils import get_column_letter
from django.shortcuts import get_object_or_404
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


def export_gantt_chart_with_order(project_id, order_by, filename_prefix):
    project = get_object_or_404(Project, id=project_id)
    
    tasks = Task.objects.filter(
        project=project
    ).select_related(
        'worker'
    ).order_by(*order_by)
    
    if not tasks.exists():
        raise NoTaskException("В этом проекте отсутствуют операции!")
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Диаграмма Ганта - {project.name}"
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')
    
    headers = ['Операция', 'Исполнитель', 'Старт (сек)', 'Финиш (сек)', 'Длительность (сек)']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
    
    row = 2
    max_time = 0
    
    for task in tasks:
        ws.cell(row=row, column=1, value=task.name)
        ws.cell(row=row, column=1).border = border
        
        ws.cell(row=row, column=2, value=task.worker.name)
        ws.cell(row=row, column=2).border = border
        
        start_time = task.start_time or 0
        ws.cell(row=row, column=3, value=start_time)
        ws.cell(row=row, column=3).border = border
        ws.cell(row=row, column=3).alignment = center_align
        
        finish_time = task.finish_time or (start_time + task.duration)
        ws.cell(row=row, column=4, value=finish_time)
        ws.cell(row=row, column=4).border = border
        ws.cell(row=row, column=4).alignment = center_align
        
        duration = task.duration or (finish_time - start_time)
        ws.cell(row=row, column=5, value=duration)
        ws.cell(row=row, column=5).border = border
        ws.cell(row=row, column=5).alignment = center_align
        
        if finish_time > max_time:
            max_time = finish_time
            
        row += 1
    
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    create_gantt_diagram(ws, tasks, max_time, row, border)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{filename_prefix}_{project_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response


def export_gantt_chart(project_id):
    return export_gantt_chart_with_order(
        project_id,
        ['worker__name', 'start_time'],
        'gantt_chart_project'
    )


def export_cyclogram_exact(project_id):
    return export_gantt_chart_with_order(
        project_id,
        ['position', 'id'],
        'cyclogram_project'
    )


def create_gantt_diagram(ws, tasks, max_time, data_start_row, border):
    diagram_start_col = 7

    time_scale_row = 1
    
    ws.cell(row=time_scale_row, column=diagram_start_col - 1, value="Время (сек):")
    ws.cell(row=time_scale_row, column=diagram_start_col - 1).font = Font(bold=True)
    ws.cell(row=time_scale_row, column=diagram_start_col - 1).alignment = Alignment(horizontal='right')
    
    for second in range(0, max_time + 11):
        col = diagram_start_col + second
        cell = ws.cell(row=time_scale_row, column=col, value=second)
        cell.font = Font(size=8, bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        cell.border = border
        
        ws.column_dimensions[get_column_letter(col)].width = 3.0

    current_row = 2

    for task in tasks:
        start_time = task.start_time or 0
        duration = task.duration or 0
        finish_time = task.finish_time or (start_time + duration)
        try:
            color = task.color if task.color and task.color.startswith('#') else '#3498db'
            fill = PatternFill(start_color=color.replace('#', ''), 
                             end_color=color.replace('#', ''), 
                             fill_type="solid")
        except:
            fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        
        for second in range(start_time, finish_time):
            col = diagram_start_col + second
            if col <= diagram_start_col + max_time + 10:
                cell = ws.cell(row=current_row, column=col)
                cell.fill = fill
                cell.border = border
        
        current_row += 1
    
    ws.row_dimensions[2].height = 20
    for row in range(3, current_row):
        ws.row_dimensions[row].height = 20
